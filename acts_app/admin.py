from django.contrib import admin
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook

from .models import Municipality, BuildingType, Act, DamageType, Damage, ActSign, Address
from django.contrib import messages
import pandas as pd
from django.http import HttpResponse
from rangefilter.filters import DateRangeFilter


@admin.register(Municipality)
class MunicipalityAdmin(admin.ModelAdmin):
    list_display = ['name']


@admin.register(BuildingType)
class BuildingTypeAdmin(admin.ModelAdmin):
    list_display = ['name', 'is_victim']


@admin.register(Address)
class AddressAdmin(admin.ModelAdmin):
    list_display = ['name', 'fias_id']


class DamageInline(admin.TabularInline):
    model = Damage
    extra = 1


@admin.register(Act)
class ActAdmin(admin.ModelAdmin):
    list_display = ['number', 'created_at', 'employee', 'victim', 'municipality', 'address', 'building_type']
    list_filter = ['created_at', ('created_at', DateRangeFilter)]
    inlines = [DamageInline]

    actions = ['export_acts_to_excel']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs

        if request.user.is_staff:
            if not request.user.municipality:
                self.message_user(request, "У вас не указан муниципалитет.", level=messages.WARNING)
                return qs.none()

            user_municipality = request.user.municipality
            filtered_qs = qs.filter(municipality=user_municipality) | qs.filter(employee=request.user)

            if not filtered_qs.exists():
                self.message_user(request, "Актов нет.", level=messages.INFO)
                return qs.none()

            return filtered_qs

        self.message_user(request, "У вас нет доступа к актам.", level=messages.ERROR)
        return qs.none()

    def export_acts_to_excel(self, request, queryset):
        if not queryset.exists():
            self.message_user(request, "Нет актов для экспорта.", level=messages.ERROR)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Акты"

        headers = [
            'Номер', 'Дата создания', 'Сотрудник', 'Пострадавший',
            'Муниципалитет', 'Адрес', 'Тип постройки', 'Время подписания',
            'Тип повреждения', 'Количество повреждений', 'Примечание'
        ]
        ws.append(headers)

        for act in queryset:
            damages = act.damages.all()
            row_start = ws.max_row + 1
            for damage in damages:
                row = [
                    act.number,
                    act.created_at.strftime("%d.%m.%Y %H:%M"),
                    str(act.employee),
                    str(act.victim) if act.victim else '',
                    str(act.municipality),
                    str(act.address),
                    str(act.building_type),
                    act.signed_at.strftime("%d.%m.%Y %H:%M") if act.signed_at else '',
                    str(damage.damage_type) if damage.damage_type else '',
                    damage.count if damage.count else '',
                    damage.note if damage.note else '',
                ]
                ws.append(row)

            if damages.exists():
                for col in range(1, 9):
                    ws.merge_cells(
                        start_row=row_start,
                        start_column=col,
                        end_row=ws.max_row,
                        end_column=col
                    )
                    ws.cell(row=row_start, column=col).alignment = Alignment(vertical='center', horizontal='center')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=acts.xlsx'

        wb.save(response)

        return response

    export_acts_to_excel.short_description = "Экспортировать в Excel"

    export_acts_to_excel.short_description = "Экспортировать в Excel"


@admin.register(DamageType)
class DamageTypeAdmin(admin.ModelAdmin):
    list_display = ['name']


@admin.register(Damage)
class DamageAdmin(admin.ModelAdmin):
    list_display = ['act', 'damage_type', 'count', 'note']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs

        if request.user.is_staff:
            if not request.user.municipality:
                self.message_user(request, "У вас не указан муниципалитет.", level=messages.WARNING)
                return qs.none()

            user_municipality = request.user.municipality
            filtered_qs = qs.filter(act__municipality=user_municipality) | qs.filter(act__employee=request.user)

            if not filtered_qs.exists():
                self.message_user(request, "Актов нет.", level=messages.INFO)
                return qs.none()

            return filtered_qs

        self.message_user(request, "У вас нет доступа к актам.", level=messages.ERROR)
        return qs.none()


@admin.register(ActSign)
class SignCodeAdmin(admin.ModelAdmin):
    list_display = ['act', 'code', 'upd_at', 'is_expired']
    readonly_fields = ['code', 'upd_at', 'is_expired']
